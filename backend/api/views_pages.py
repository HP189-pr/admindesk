import os
import uuid
import logging
import pandas as pd
from django.conf import settings
from django.utils import timezone
from django.db import transaction
from django.core.cache import cache
from django.db.models import Q
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from openpyxl import load_workbook
from .models import Institute, MainBranch, SubBranch, Enrollment
from .serializers import (
    InstituteSerializer, MainBranchSerializer, SubBranchSerializer, EnrollmentSerializer,
    VerificationSerializer, EcaResendSerializer, ResubmitSerializer, AssignFinalSerializer, Verification
)

logger = logging.getLogger(__name__)

# ✅ Institute ViewSet
class InstituteViewSet(viewsets.ModelViewSet):
    queryset = Institute.objects.all()
    serializer_class = InstituteSerializer

# ✅ Main Branch ViewSet
class MainBranchViewSet(viewsets.ModelViewSet):
    queryset = MainBranch.objects.all()
    serializer_class = MainBranchSerializer

# ✅ Sub Branch ViewSet
class SubBranchViewSet(viewsets.ModelViewSet):
    queryset = SubBranch.objects.all()
    serializer_class = SubBranchSerializer

# ✅ Enrollment ViewSet
class EnrollmentViewSet(viewsets.ModelViewSet):
    queryset = Enrollment.objects.all()
    serializer_class = EnrollmentSerializer
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['POST'], url_path='init-upload')
    def init_upload(self, request):
        """Initialize large file upload session."""
        try:
            uploaded_file = request.FILES.get('file')
            if not uploaded_file:
                return Response({"error": "No file provided in request"}, status=status.HTTP_400_BAD_REQUEST)

            # Validate file type
            valid_extensions = ['.xlsx', '.xls']
            file_ext = os.path.splitext(uploaded_file.name)[1].lower()
            if file_ext not in valid_extensions:
                return Response(
                    {"error": "Invalid file type. Only Excel files (.xlsx, .xls) are allowed"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Save file temporarily
            temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_uploads')
            os.makedirs(temp_dir, exist_ok=True)
            temp_filename = f"upload_{uuid.uuid4()}{file_ext}"
            temp_path = os.path.join(temp_dir, temp_filename)

            with open(temp_path, 'wb+') as destination:
                for chunk in uploaded_file.chunks():
                    destination.write(chunk)

            # Create upload session
            session_id = str(uuid.uuid4())
            cache_data = {
                'file_path': temp_path,
                'original_filename': uploaded_file.name,
                'user_id': request.user.id,
                'progress': 0,
                'stage': 'initialized',
                'created_at': timezone.now().isoformat()
            }
            cache.set(f"upload_{session_id}", cache_data, timeout=3600)

            return Response({
                'session_id': session_id,
                'original_filename': uploaded_file.name,
                'message': 'Upload initialized successfully'
            })

        except Exception as e:
            logger.error(f"Upload initialization failed: {str(e)}", exc_info=True)
            return Response(
                {"error": "Internal server error during upload initialization", "details": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['POST'], url_path='get-sheets')
    def get_sheets(self, request):
        """Retrieve sheet names from the uploaded Excel file."""
        session_id = request.data.get('session_id')
        if not session_id:
            return Response({"error": "Session ID required"}, status=status.HTTP_400_BAD_REQUEST)

        upload_data = cache.get(f"upload_{session_id}")
        if not upload_data:
            return Response({"error": "Invalid session"}, status=status.HTTP_404_NOT_FOUND)

        try:
            with load_workbook(upload_data['file_path'], read_only=True) as wb:
                return Response({'sheets': wb.sheetnames})
        except Exception as e:
            logger.error(f"Error reading sheets: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['POST'], url_path='get-columns')
    def get_columns(self, request):
        """Retrieve column headers from a specific sheet."""
        session_id = request.data.get('session_id')
        sheet_name = request.data.get('sheet_name')

        if not all([session_id, sheet_name]):
            return Response({"error": "Session ID and sheet name required"}, status=status.HTTP_400_BAD_REQUEST)

        upload_data = cache.get(f"upload_{session_id}")
        if not upload_data:
            return Response({"error": "Invalid session"}, status=status.HTTP_404_NOT_FOUND)

        try:
            df = pd.read_excel(upload_data['file_path'], sheet_name=sheet_name, nrows=1, engine='openpyxl')
            return Response({'columns': list(df.columns)})
        except Exception as e:
            logger.error(f"Error reading columns: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['POST'], url_path='process-chunk')
    def process_chunk(self, request):
        """Process data in chunks with progress tracking."""
        session_id = request.data.get('session_id')
        sheet_name = request.data.get('sheet_name')
        column_mapping = request.data.get('column_mapping', {})
        chunk_size = 1000

        if not all([session_id, sheet_name, column_mapping]):
            return Response({"error": "Missing required parameters"}, status=status.HTTP_400_BAD_REQUEST)

        upload_data = cache.get(f"upload_{session_id}")
        if not upload_data:
            return Response({"error": "Invalid session"}, status=status.HTTP_404_NOT_FOUND)

        try:
            results = {'total_processed': 0, 'success_count': 0, 'errors': []}

            for chunk in pd.read_excel(
                upload_data['file_path'], sheet_name=sheet_name, chunksize=chunk_size, engine='openpyxl'
            ):
                chunk = chunk.rename(columns={v: k for k, v in column_mapping.items()})
                chunk_results = self._process_chunk_data(chunk, request.user)

                results['total_processed'] += len(chunk)
                results['success_count'] += chunk_results['success_count']
                results['errors'].extend(chunk_results['errors'])

                cache.set(f"upload_{session_id}", {**upload_data, 'progress': results['total_processed']}, timeout=3600)

            return Response(results)
        except Exception as e:
            logger.error(f"Chunk processing failed: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _process_chunk_data(self, chunk, user):
        """Process a single chunk of data."""
        institutes = {str(i.id): i for i in Institute.objects.all()}
        maincourses = {str(m.id): m for m in MainBranch.objects.all()}
        subcourses = {str(s.id): s for s in SubBranch.objects.all()}

        successes = 0
        errors = []

        with transaction.atomic():
            for _, row in chunk.iterrows():
                try:
                    enrollment_no = str(row.get('enrollment_no', '')).strip()
                    if not enrollment_no:
                        raise ValueError("Missing enrollment_no")

                    if Enrollment.objects.filter(enrollment_no=enrollment_no).exists():
                        raise ValueError("Duplicate enrollment_no")

                    Enrollment.objects.create(
                        enrollment_no=enrollment_no,
                        student_name=row.get('student_name', ''),
                        institute=institutes.get(str(row.get('institute_id'))),
                        batch=int(row.get('batch', 0)),
                        admission_date=row.get('admission_date'),
                        subcourse=subcourses.get(str(row.get('subcourse_id'))),
                        maincourse=maincourses.get(str(row.get('maincourse_id'))),
                        updated_by=user
                    )
                    successes += 1
                except Exception as e:
                    errors.append({'row': _, 'error': str(e), 'data': dict(row.dropna())})

        return {'success_count': successes, 'errors': errors}
class VerificationViewSet(viewsets.ModelViewSet):
    """
    Endpoints:
      GET    /api/verification/              (list with ?q=&limit=)
      POST   /api/verification/              (create)
      GET    /api/verification/{id}/         (retrieve)
      PATCH  /api/verification/{id}/         (partial update)
      PUT    /api/verification/{id}/         (update)
      POST   /api/verification/{id}/eca-resend/
      POST   /api/verification/{id}/resubmit/
      POST   /api/verification/{id}/assign-final/
    """
    serializer_class = VerificationSerializer
    queryset = (Verification.objects
                .select_related("enrollment", "second_enrollment", "updatedby")
                .order_by("-id"))
    permission_classes = [IsAuthenticated]  # or your custom permission

    def get_queryset(self):
        qs = super().get_queryset()
        q = self.request.query_params.get("q")
        if q:
            qs = qs.filter(
                Q(student_name__icontains=q) |
                Q(final_no__icontains=q) |
                Q(pay_rec_no__icontains=q) |
                Q(eca_ref_no__icontains=q) |
                Q(enrollment__enrollment_no__icontains=q) |
                Q(second_enrollment__enrollment_no__icontains=q)
            )
        limit = self.request.query_params.get("limit")
        if limit:
            try:
                n = int(limit)
                if n > 0:
                    qs = qs[:n]
            except ValueError:
                pass
        return qs

    def perform_create(self, serializer):
        # updatedby audit
        user = self.request.user if (self.request and self.request.user.is_authenticated) else None
        serializer.save(updatedby=user)

    def perform_update(self, serializer):
        user = self.request.user if (self.request and self.request.user.is_authenticated) else None
        serializer.save(updatedby=user)

    # --- Custom Actions ---

    @action(detail=True, methods=["post"], url_path="eca-resend")
    def eca_resend(self, request, pk=None):
        """
        Append a RESEND to ECA history, update counters & mail status.
        Body: { "to_email": "...", "notes": "..." }
        """
        instance: Verification = self.get_object()
        ser = EcaResendSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        to_email = ser.validated_data["to_email"]
        notes = ser.validated_data.get("notes", "")

        # Push to history (helper on model)
        instance.eca_push_history(action="RESEND", to_email=to_email, notes=notes, mark_sent=True)

        return Response(VerificationSerializer(instance, context={"request": request}).data)

    @action(detail=True, methods=["post"], url_path="resubmit")
    def resubmit(self, request, pk=None):
        """
        Mark a resubmission: stamp last_resubmit_* and set status back to IN_PROGRESS.
        Body: { "status_note": "..." } (optional)
        """
        instance: Verification = self.get_object()
        ser = ResubmitSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        note = ser.validated_data.get("status_note")
        instance.record_resubmit(status_note=note)

        return Response(VerificationSerializer(instance, context={"request": request}).data)

    @action(detail=True, methods=["post"], url_path="assign-final")
    def assign_final(self, request, pk=None):
        """
        Assign/overwrite a final number on the SAME row and mark DONE.
        Body: { "final_no": "TR-2025-000123" }
        (If you ever need 'new number creates new row', we can add a different action.)
        """
        instance: Verification = self.get_object()
        ser = AssignFinalSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        final_no = ser.validated_data["final_no"]
        instance.final_no = final_no
        instance.status = VerificationStatus.DONE
        instance.full_clean()   # apply model-level validation
        instance.save(update_fields=["final_no", "status", "updatedat"])

        return Response(VerificationSerializer(instance, context={"request": request}).data)

